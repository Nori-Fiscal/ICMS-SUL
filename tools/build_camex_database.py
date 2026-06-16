import csv
import hashlib
import json
import re
import zipfile
from datetime import date, datetime
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple
from xml.etree import ElementTree as ET

import openpyxl
import requests


ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data"
SOURCE_DIR = DATA_DIR / "camex_sources"
OUT_CSV = DATA_DIR / "camex_ncm_database.csv"
OUT_SUMMARY = DATA_DIR / "camex_ncm_summary.json"
OUT_MANIFEST = SOURCE_DIR / "source_manifest.json"


SOURCES = [
    {
        "file": "res_272_anexos_i_x.xlsx",
        "url": "https://www.gov.br/mdic/pt-br/assuntos/camex/estrategia-comercial/arquivos-listas/01-06-2026-anexos-i-a-x-resolucao-gecex-272-21-1.xlsx",
        "updated_at": "2026-06-01",
    },
    {
        "file": "lessin_consolidada.xlsx",
        "url": "https://www.gov.br/mdic/pt-br/assuntos/camex/estrategia-comercial/arquivos-listas/lessin.xlsx",
        "updated_at": "2026-06-08",
    },
    {
        "file": "cotas_camex_vigentes.ods",
        "url": "https://www.gov.br/siscomex/pt-br/informacoes/importacaoo/2026.05.21_CotasCAMEXvigentes.ods",
        "updated_at": "2026-05-21",
    },
    {
        "file": "cotas_camex_acompanhamento.ods",
        "url": "https://www.gov.br/siscomex/pt-br/informacoes/importacaoo/2026.05.30_Tabeladeacompanhamentodascotasdeimportao.ods",
        "updated_at": "2026-05-30",
    },
    {
        "file": "modelos_lpco_cotas.ods",
        "url": "https://www.gov.br/siscomex/pt-br/informacoes/importacaoo/2026.05.21_ModelosLPCO_Decex_Cotas.ods",
        "updated_at": "2026-05-21",
    },
    {
        "file": "ex_tarifarios_bk_bit_vigentes.xlsx",
        "url": "https://www.gov.br/mdic/pt-br/assuntos/sdic/ex-tarifario/pda-de-ex-tarifarios/ex-tarifarios-vigentes.xlsx",
        "updated_at": "2026-06-08",
    },
    {
        "file": "ex_tarifarios_bk_auto_vigentes.xlsx",
        "url": "https://www.gov.br/mdic/pt-br/assuntos/sdic/setor-automotivo/bk-autopropulsado/ExTarifariosVigentes_BKAUTO2025dezRes841.xlsx",
        "updated_at": "2025-12-01",
    },
]


CSV_COLUMNS = [
    "ncm",
    "ncm_formatado",
    "match_tipo",
    "prefix_len",
    "ex",
    "lista",
    "categoria",
    "descricao",
    "aliquota",
    "quota",
    "unidade_quota",
    "inicio_vigencia",
    "fim_vigencia",
    "ato_legal",
    "portaria_secex",
    "modelo_lpco",
    "observacao",
    "fonte_arquivo",
    "fonte_url",
    "fonte_atualizacao",
]


def clean(value) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.date().isoformat() if isinstance(value, datetime) else value.isoformat()
    text = str(value).replace("\xa0", " ").replace("\n", " ").strip()
    text = re.sub(r"\s+", " ", text)
    return "" if text in {"-", "--", "nan", "None"} else text


def parse_date(value) -> str:
    text = clean(value)
    if not text:
        return ""
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}(?: .*)?", text):
        return text[:10]
    match = re.fullmatch(r"(\d{1,2})/(\d{1,2})/(\d{4})", text)
    if match:
        day, month, year = match.groups()
        return f"{year}-{int(month):02d}-{int(day):02d}"
    return text


def split_period(value) -> Tuple[str, str]:
    text = clean(value)
    if not text:
        return "", ""
    parts = re.split(r"\s+a\s+", text, maxsplit=1, flags=re.IGNORECASE)
    if len(parts) == 2:
        return parse_date(parts[0]), parse_date(parts[1])
    return "", ""


def only_digits(value) -> str:
    return re.sub(r"\D", "", clean(value))


def format_ncm(ncm: str) -> str:
    if len(ncm) == 8:
        return f"{ncm[:4]}.{ncm[4:6]}.{ncm[6:8]}"
    if len(ncm) == 6:
        return f"{ncm[:4]}.{ncm[4:6]}"
    return ncm


def extract_ncms(value, allow_prefix: bool = False) -> List[str]:
    text = clean(value)
    if not text:
        return []
    exact = [re.sub(r"\D", "", m) for m in re.findall(r"\b\d{4}\.\d{2}\.\d{2}\b", text)]
    exact += [m for m in re.findall(r"(?<!\d)(\d{8})(?!\d)", text)]
    if exact:
        return sorted(set(exact))
    if allow_prefix:
        prefixes = [re.sub(r"\D", "", m) for m in re.findall(r"\b\d{4}\.\d{2}\b", text)]
        prefixes += [m for m in re.findall(r"(?<!\d)(\d{6})(?!\d)", text)]
        return sorted(set(prefixes))
    return []


def normalize_ex(value) -> str:
    text = clean(value)
    if not text:
        return ""
    nums = re.findall(r"\d{1,3}", text)
    if not nums:
        return ""
    return nums[0].zfill(3)


def extract_exs(*values) -> List[str]:
    joined = " ".join(clean(v) for v in values if clean(v))
    if not joined:
        return [""]
    lower = joined.lower()
    if "ex" in lower:
        first = lower.find("ex")
        nums = re.findall(r"\b\d{1,3}\b", joined[first:])
        return sorted({n.zfill(3) for n in nums}) or [""]
    if re.fullmatch(r"\d{1,3}", joined):
        return [joined.zfill(3)]
    return [""]


def source_meta(filename: str) -> Dict[str, str]:
    for src in SOURCES:
        if src["file"] == filename:
            return src
    return {"file": filename, "url": "", "updated_at": ""}


def row_dict(headers: Sequence[str], row: Sequence) -> Dict[str, str]:
    out: Dict[str, str] = {}
    for idx, header in enumerate(headers):
        if not header:
            continue
        out[clean(header).lower()] = clean(row[idx] if idx < len(row) else "")
    return out


def add_record(
    rows: List[Dict[str, str]],
    *,
    ncm: str,
    ex: str = "",
    lista: str,
    categoria: str,
    descricao: str = "",
    aliquota: str = "",
    quota: str = "",
    unidade_quota: str = "",
    inicio_vigencia: str = "",
    fim_vigencia: str = "",
    ato_legal: str = "",
    portaria_secex: str = "",
    modelo_lpco: str = "",
    observacao: str = "",
    fonte_arquivo: str,
    match_tipo: Optional[str] = None,
) -> None:
    meta = source_meta(fonte_arquivo)
    ncm = only_digits(ncm)
    if len(ncm) not in {6, 8}:
        return
    if match_tipo is None:
        match_tipo = "prefixo" if len(ncm) == 6 else "exato"
    rows.append(
        {
            "ncm": ncm,
            "ncm_formatado": format_ncm(ncm),
            "match_tipo": match_tipo,
            "prefix_len": str(len(ncm)),
            "ex": normalize_ex(ex),
            "lista": clean(lista),
            "categoria": clean(categoria),
            "descricao": clean(descricao),
            "aliquota": clean(aliquota),
            "quota": clean(quota),
            "unidade_quota": clean(unidade_quota),
            "inicio_vigencia": parse_date(inicio_vigencia),
            "fim_vigencia": parse_date(fim_vigencia),
            "ato_legal": clean(ato_legal),
            "portaria_secex": clean(portaria_secex),
            "modelo_lpco": clean(modelo_lpco),
            "observacao": clean(observacao),
            "fonte_arquivo": fonte_arquivo,
            "fonte_url": meta.get("url", ""),
            "fonte_atualizacao": meta.get("updated_at", ""),
        }
    )


def iter_excel_dicts(path: Path, sheet_name: str, header_row: int) -> Iterable[Dict[str, str]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    headers = [clean(c.value) for c in next(ws.iter_rows(min_row=header_row, max_row=header_row))]
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        data = row_dict(headers, row)
        if any(data.values()):
            yield data


def read_ods(path: Path) -> Dict[str, List[List[str]]]:
    ns = {
        "table": "urn:oasis:names:tc:opendocument:xmlns:table:1.0",
        "text": "urn:oasis:names:tc:opendocument:xmlns:text:1.0",
    }
    table_ns = "{urn:oasis:names:tc:opendocument:xmlns:table:1.0}"
    sheets: Dict[str, List[List[str]]] = {}
    with zipfile.ZipFile(path) as archive:
        root = ET.fromstring(archive.read("content.xml"))
    for table in root.findall(".//table:table", ns):
        name = table.attrib.get(table_ns + "name", "Planilha")
        rows: List[List[str]] = []
        for tr in table.findall("table:table-row", ns):
            row_repeat = int(tr.attrib.get(table_ns + "number-rows-repeated", "1"))
            cells: List[str] = []
            for cell in tr.findall("table:table-cell", ns):
                col_repeat = int(cell.attrib.get(table_ns + "number-columns-repeated", "1"))
                texts = []
                for p in cell.findall(".//text:p", ns):
                    texts.append("".join(p.itertext()))
                value = clean(" ".join(texts))
                cells.extend([value] * min(col_repeat, 200))
            for _ in range(min(row_repeat, 5000)):
                rows.append(cells)
        sheets[name] = rows
    return sheets


def download_sources() -> List[Dict[str, str]]:
    SOURCE_DIR.mkdir(parents=True, exist_ok=True)
    headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}
    manifest = []
    for src in SOURCES:
        path = SOURCE_DIR / src["file"]
        response = requests.get(src["url"], headers=headers, timeout=90)
        response.raise_for_status()
        content = response.content
        path.write_bytes(content)
        manifest.append(
            {
                **src,
                "downloaded_at": datetime.now().isoformat(timespec="seconds"),
                "bytes": len(content),
                "sha256": hashlib.sha256(content).hexdigest(),
                "content_type": response.headers.get("content-type", ""),
            }
        )
    OUT_MANIFEST.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    return manifest


def parse_res_272(rows: List[Dict[str, str]]) -> None:
    filename = "res_272_anexos_i_x.xlsx"
    path = SOURCE_DIR / filename

    # Anexo I is the full TEC and would alert almost every valid NCM. It is deliberately excluded.
    for data in iter_excel_dicts(path, "Anexo II - Diferentes da TEC", 3):
        for ncm in extract_ncms(data.get("ncm")):
            add_record(
                rows,
                ncm=ncm,
                lista="Anexo II - Tarifas brasileiras diferentes da TEC",
                categoria="Resolucao GECEX 272/2021",
                descricao=data.get("descrição") or data.get("descricao"),
                aliquota=data.get("alíquota aplicada (%)"),
                ato_legal=data.get("atos de inclusão"),
                observacao=f"TEC: {data.get('tec (%)', '')}; BIT/BK: {data.get('bit/bk', '')}; Fundamento: {data.get('fundamentação da alíquota aplicada', '')}",
                fonte_arquivo=filename,
            )

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Anexo III - Setor Aeronáutico"]
    for row in ws.iter_rows(min_row=5, values_only=True):
        for cell in row:
            for ncm in extract_ncms(cell, allow_prefix=True):
                add_record(
                    rows,
                    ncm=ncm,
                    lista="Anexo III - Setor Aeronautico",
                    categoria="Resolucao GECEX 272/2021",
                    descricao="Regra de tributacao do Mercosul para produtos do setor aeronautico",
                    fonte_arquivo=filename,
                    match_tipo="prefixo",
                )

    sheet_headers = {
        "Anexo IV - Desabastecimento": 5,
        "Anexo V - LETEC": 5,
        "Anexo VI - LEBITBK": 4,
        "Anexo VIII - Concessões OMC": 4,
        "Anexo IX - DCC": 4,
        "Anexo X - Automotivos ACE-14": 5,
    }
    for sheet, header_row in sheet_headers.items():
        for data in iter_excel_dicts(path, sheet, header_row):
            ncm_values = extract_ncms(data.get("ncm"))
            if not ncm_values:
                continue
            for ncm in ncm_values:
                for ex in extract_exs(data.get("nº ex"), data.get("nº ex")):
                    add_record(
                        rows,
                        ncm=ncm,
                        ex=ex,
                        lista=sheet,
                        categoria="Resolucao GECEX 272/2021",
                        descricao=data.get("descrição") or data.get("descricao"),
                        aliquota=data.get("alíquota (%)"),
                        quota=data.get("quota"),
                        unidade_quota=data.get("unidade da quota") or data.get("unidade quota"),
                        inicio_vigencia=data.get("início de vigência") or data.get("início da vigência"),
                        fim_vigencia=data.get("término de vigência"),
                        ato_legal=data.get("ato de inclusão") or data.get("ato de inclusao"),
                        observacao=data.get("observações") or data.get("observação") or data.get("enquadramento anexo res. gmc 49/19"),
                        fonte_arquivo=filename,
                    )


def parse_lessin(rows: List[Dict[str, str]]) -> None:
    filename = "lessin_consolidada.xlsx"
    path = SOURCE_DIR / filename
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    for row in ws.iter_rows(min_row=21, values_only=True):
        values = list(row) + [""] * 17
        ncm = values[1]
        if not extract_ncms(ncm):
            continue
        grupo = clean(values[13])
        fundamento = clean(values[14])
        atos = "; ".join(clean(v) for v in values[10:13] if clean(v))
        for ncm_value in extract_ncms(ncm):
            for ex in extract_exs(values[2]):
                add_record(
                    rows,
                    ncm=ncm_value,
                    ex=ex,
                    lista=f"Lessin - {grupo}" if grupo else "Lessin",
                    categoria="Lista de Bens sem Similar Nacional",
                    descricao=values[3],
                    aliquota=values[4],
                    quota=values[6],
                    unidade_quota=values[7],
                    inicio_vigencia=values[8],
                    fim_vigencia=values[9],
                    ato_legal=atos,
                    observacao=f"{fundamento}; Quota: {clean(values[5])}",
                    fonte_arquivo=filename,
                )


def add_cota_group(rows: List[Dict[str, str]], group: Optional[Dict], ex_rows: List[str]) -> None:
    if not group:
        return
    ex_values = ex_rows or [""]
    for ncm in group["ncms"]:
        for ex_text in ex_values:
            exs = extract_exs(ex_text)
            descricao = ex_text if ex_text else group["descricao"]
            for ex in exs:
                add_record(
                    rows,
                    ncm=ncm,
                    ex=ex,
                    lista="Cotas CAMEX vigentes",
                    categoria="Cota de importacao",
                    descricao=descricao,
                    aliquota=group["aliquota"],
                    quota=group["quota"],
                    inicio_vigencia=group["inicio"],
                    fim_vigencia=group["fim"],
                    ato_legal=group["resolucao"],
                    portaria_secex=group["portaria"],
                    observacao=group["observacao"],
                    fonte_arquivo="cotas_camex_vigentes.ods",
                )


def parse_cotas_vigentes(rows: List[Dict[str, str]]) -> None:
    sheets = read_ods(SOURCE_DIR / "cotas_camex_vigentes.ods")
    sheet_rows = next(iter(sheets.values()))
    group = None
    ex_rows: List[str] = []
    for row in sheet_rows[2:]:
        first = clean(row[0] if row else "")
        if not first:
            continue
        ncms = extract_ncms(first)
        if ncms:
            add_cota_group(rows, group, ex_rows)
            inicio, fim = split_period(row[7] if len(row) > 7 else "")
            group = {
                "ncms": ncms,
                "descricao": row[1] if len(row) > 1 else "",
                "aliquota": row[2] if len(row) > 2 else "",
                "quota": row[3] if len(row) > 3 else "",
                "inicio": inicio,
                "fim": fim,
                "resolucao": row[8] if len(row) > 8 else "",
                "portaria": row[9] if len(row) > 9 else "",
                "observacao": "; ".join(clean(v) for v in row[4:7] if clean(v)),
            }
            ex_rows = []
        elif first.lower().startswith("ex ") and group:
            ex_rows.append(first)
    add_cota_group(rows, group, ex_rows)


def parse_cotas_acompanhamento(rows: List[Dict[str, str]]) -> None:
    sheets = read_ods(SOURCE_DIR / "cotas_camex_acompanhamento.ods")
    sheet_rows = next(iter(sheets.values()))
    headers = [clean(v).lower() for v in sheet_rows[2]]
    for raw in sheet_rows[3:]:
        data = row_dict(headers, raw)
        for ncm in extract_ncms(data.get("ncm")):
            for ex in extract_exs(data.get("ncm")):
                add_record(
                    rows,
                    ncm=ncm,
                    ex=ex,
                    lista="Acompanhamento das cotas de importacao",
                    categoria="Cota de importacao",
                    quota=data.get("cota concedida"),
                    unidade_quota=data.get("unidade de medida da cota"),
                    inicio_vigencia=data.get("início vigência"),
                    fim_vigencia=data.get("fim vigência"),
                    ato_legal=data.get("resolução gecex"),
                    portaria_secex=data.get("portaria secex"),
                    observacao=f"Cota consumida: {data.get('cota consumida', '')}; Percentual: {data.get('percentual de consumo (g)/(f)', '')}",
                    fonte_arquivo="cotas_camex_acompanhamento.ods",
                )


def parse_modelos_lpco(rows: List[Dict[str, str]]) -> None:
    sheets = read_ods(SOURCE_DIR / "modelos_lpco_cotas.ods")
    sheet_rows = next(iter(sheets.values()))
    headers = [clean(v).lower() for v in sheet_rows[0]]
    for raw in sheet_rows[1:]:
        data = row_dict(headers, raw)
        ncm_values = extract_ncms(data.get("ncm"))
        if not ncm_values:
            continue
        ex_values = extract_exs(data.get("produto")) if data.get("ex-tarifário (sim ou não)", "").lower() == "sim" else [""]
        for ncm in ncm_values:
            for ex in ex_values:
                add_record(
                    rows,
                    ncm=ncm,
                    ex=ex,
                    lista="Modelos LPCO Decex - Cotas",
                    categoria=data.get("tipo de cota") or "Cota de importacao",
                    descricao=data.get("produto"),
                    unidade_quota=data.get("unidade de controle da cota"),
                    inicio_vigencia=data.get("início vigência"),
                    fim_vigencia=data.get("fim vigência"),
                    ato_legal=data.get("resolução gecex"),
                    portaria_secex=data.get("portaria secex"),
                    modelo_lpco=data.get("modelo lpco"),
                    observacao=data.get("critério distribuição"),
                    fonte_arquivo="modelos_lpco_cotas.ods",
                )


def parse_ex_tarifarios(rows: List[Dict[str, str]]) -> None:
    filename = "ex_tarifarios_bk_bit_vigentes.xlsx"
    for data in iter_excel_dicts(SOURCE_DIR / filename, "Planilha1", 1):
        for ncm in extract_ncms(data.get("ncm")):
            add_record(
                rows,
                ncm=ncm,
                ex=data.get("ex"),
                lista="Ex-tarifarios BK/BIT vigentes",
                categoria="Ex-tarifario",
                descricao=data.get("descrição"),
                inicio_vigencia=data.get("início da vigência"),
                fim_vigencia=data.get("vigência"),
                ato_legal=data.get("ato legal"),
                observacao=f"Anexo: {data.get('anexo', '')}; Inclusao: {data.get('resolução de inclusão', '')}",
                fonte_arquivo=filename,
            )

    filename = "ex_tarifarios_bk_auto_vigentes.xlsx"
    for data in iter_excel_dicts(SOURCE_DIR / filename, "Ex-Tarifários BK Autopropulsado", 1):
        for ncm in extract_ncms(data.get("ncm")):
            add_record(
                rows,
                ncm=ncm,
                ex=data.get("ex"),
                lista="Ex-tarifarios BK Autopropulsado",
                categoria="Ex-tarifario",
                descricao=data.get("descrição"),
                fim_vigencia=data.get("fim da vigência"),
                ato_legal=data.get("publicação"),
                observacao=f"Chave: {data.get('chave', '')}",
                fonte_arquivo=filename,
            )


def dedupe(rows: List[Dict[str, str]]) -> List[Dict[str, str]]:
    seen = set()
    out = []
    for row in rows:
        key = tuple(row.get(col, "") for col in CSV_COLUMNS)
        if key in seen:
            continue
        seen.add(key)
        out.append(row)
    return out


def build_database() -> List[Dict[str, str]]:
    rows: List[Dict[str, str]] = []
    parse_res_272(rows)
    parse_lessin(rows)
    parse_cotas_vigentes(rows)
    parse_cotas_acompanhamento(rows)
    parse_modelos_lpco(rows)
    parse_ex_tarifarios(rows)
    return dedupe(rows)


def write_outputs(rows: List[Dict[str, str]]) -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    with OUT_CSV.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=CSV_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)

    summary = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "rows": len(rows),
        "unique_ncm_exact": len({r["ncm"] for r in rows if r["match_tipo"] == "exato"}),
        "unique_ncm_prefix": len({r["ncm"] for r in rows if r["match_tipo"] == "prefixo"}),
        "by_lista": {},
        "sources": SOURCES,
        "notes": [
            "Anexo I - TEC foi excluido do alerta porque contem a tarifa geral e marcaria praticamente qualquer NCM valido.",
            "Registros com NCM de 6 digitos sao tratados como prefixo para bater nos NCMs de 8 digitos do XML.",
            "Registros com EX indicam que a classificacao depende do EX; o app alerta pelo NCM e mostra o EX para conferencia.",
        ],
    }
    for row in rows:
        summary["by_lista"][row["lista"]] = summary["by_lista"].get(row["lista"], 0) + 1
    OUT_SUMMARY.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")


def main() -> None:
    download_sources()
    rows = build_database()
    write_outputs(rows)
    print(f"CSV: {OUT_CSV}")
    print(f"Linhas: {len(rows)}")
    print(f"NCMs exatos unicos: {len({r['ncm'] for r in rows if r['match_tipo'] == 'exato'})}")
    print(f"Prefixos unicos: {len({r['ncm'] for r in rows if r['match_tipo'] == 'prefixo'})}")


if __name__ == "__main__":
    main()
