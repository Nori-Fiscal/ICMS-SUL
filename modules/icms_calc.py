# -*- coding: utf-8 -*-
"""
Módulo de cálculo ICMS TDD 409 — extraído do app_icms_albema.py
"""
from __future__ import annotations

import io
import re
import unicodedata
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

URL_SAT_SC = (
    "https://sat.sef.sc.gov.br/tax.NET/RequestClientCertificate.aspx?"
    "returnUrl=https%3a%2f%2fsat.sef.sc.gov.br%2ftax.net%2f"
    "Sat.ComercioExterior.Web%2fDuimp%2fPainelDeclaracoesIcms.aspx"
)

COLUNAS_OBRIGATORIAS: Dict[str, Dict[str, Any]] = {
    "valor_total": {"rotulo": "Valor Total", "aliases": ["valor total", "vlr total", "valor mercadoria", "total"]},
    "valor_ii": {"rotulo": "Valor II", "aliases": ["valor ii", "vlr ii", "imposto importacao", "imposto de importacao"]},
    "valor_ipi": {"rotulo": "Valor IPI", "aliases": ["valor ipi", "vlr ipi", "ipi"]},
    "valor_pis": {"rotulo": "Valor PIS", "aliases": ["valor pis", "vlr pis", "pis"]},
    "valor_cofins": {"rotulo": "Valor Cofins", "aliases": ["valor cofins", "vlr cofins", "cofins"]},
    "siscomex": {"rotulo": "Siscomex", "aliases": ["siscomex", "taxa siscomex", "valor siscomex"]},
    "afrmm": {"rotulo": "AFRMM", "aliases": ["afrmm", "valor afrmm"]},
}
ALIASES_CONFERENCIA = {
    "base_icms_original": ["base icms", "vlr base icms", "base do icms"],
    "valor_icms_original": ["valor icms", "vlr icms", "icms"],
}

def normalizar_texto(valor: Any) -> str:
    texto = "" if valor is None else str(valor)
    texto = texto.strip().lower()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    texto = re.sub(r"[^a-z0-9]+", " ", texto)
    return re.sub(r"\s+", " ", texto).strip()

def contem_todas_palavras(nome_coluna_normalizado: str, alias_normalizado: str) -> bool:
    palavras_coluna = set(nome_coluna_normalizado.split())
    palavras_alias = alias_normalizado.split()
    return bool(palavras_alias) and all(p in palavras_coluna for p in palavras_alias)

def converter_numero(valor: Any) -> float:
    if valor is None or pd.isna(valor):
        return 0.0
    if isinstance(valor, bool):
        return 0.0
    if isinstance(valor, (int, float)):
        return float(valor)
    texto = str(valor).strip()
    if texto == "" or texto.lower() in {"nan", "none", "null", "na", "n/a", "-"}:
        return 0.0
    texto = texto.replace("\u00a0", "")
    texto = texto.replace("R$", "").replace("r$", "")
    texto = texto.replace(" ", "")
    neg = texto.startswith("(") and texto.endswith(")")
    if neg:
        texto = texto[1:-1]
    texto = re.sub(r"[^0-9,.\-]", "", texto)
    if texto in {"", "-", ",", ".", "-0", "-0,00", "-0.00"}:
        return 0.0
    sinal = -1 if ("-" in texto or neg) else 1
    texto = texto.replace("-", "")
    if "," in texto and "." in texto:
        if texto.rfind(",") > texto.rfind("."):
            texto = texto.replace(".", "").replace(",", ".")
        else:
            texto = texto.replace(",", "")
    elif "," in texto:
        texto = texto.replace(".", "").replace(",", ".")
    elif "." in texto:
        partes = texto.split(".")
        if len(partes) > 2:
            if len(partes[-1]) == 3:
                texto = "".join(partes)
            else:
                texto = "".join(partes[:-1]) + "." + partes[-1]
        elif len(partes) == 2 and len(partes[1]) == 3 and len(partes[0]) <= 3:
            texto = partes[0] + partes[1]
    try:
        return sinal * float(texto)
    except ValueError:
        return 0.0

def formatar_brl(valor: float) -> str:
    texto = f"R$ {valor:,.2f}"
    return texto.replace(",", "X").replace(".", ",").replace("X", ".")

def localizar_coluna(df: pd.DataFrame, aliases: List[str]) -> Optional[str]:
    col_norm = {normalizar_texto(c): c for c in df.columns}
    for alias in aliases:
        a = normalizar_texto(alias)
        if a in col_norm:
            return col_norm[a]
    for alias in aliases:
        a = normalizar_texto(alias)
        for cn, co in col_norm.items():
            if contem_todas_palavras(cn, a):
                return co
    return None

def localizar_colunas_obrigatorias(df: pd.DataFrame) -> Dict[str, Optional[str]]:
    return {c: localizar_coluna(df, cfg["aliases"]) for c, cfg in COLUNAS_OBRIGATORIAS.items()}

def localizar_colunas_conferencia(df: pd.DataFrame) -> Dict[str, Optional[str]]:
    return {c: localizar_coluna(df, aliases) for c, aliases in ALIASES_CONFERENCIA.items()}

def identificar_linhas_total(df: pd.DataFrame) -> pd.Series:
    aliases_texto = {
        "adicao": ["adicao", "adicao numero"],
        "item": ["item"],
        "codigo": ["codigo", "cod", "sku", "referencia"],
        "descricao": ["descricao", "produto", "mercadoria"],
    }
    cols = [localizar_coluna(df, a) for a in aliases_texto.values()]
    cols = [c for c in cols if c is not None]
    if not cols:
        return pd.Series(False, index=df.index)
    return df[cols].fillna("").astype(str).apply(lambda c: c.str.strip().eq("")).all(axis=1)

def ler_arquivo_excel_ou_csv(arquivo_nome: str, arquivo_bytes: bytes) -> Tuple[pd.DataFrame, str]:
    if arquivo_nome.lower().endswith(".csv"):
        for sep in [";", "\t", ","]:
            try:
                df = pd.read_csv(io.BytesIO(arquivo_bytes), sep=sep, dtype=str)
                if df.shape[1] > 1:
                    return df, "CSV"
            except Exception:
                continue
        return pd.read_csv(io.BytesIO(arquivo_bytes), dtype=str), "CSV"
    excel = pd.ExcelFile(io.BytesIO(arquivo_bytes))
    aba = st.sidebar.selectbox("Aba da planilha", excel.sheet_names, key="icms_aba")
    return pd.read_excel(io.BytesIO(arquivo_bytes), sheet_name=aba, dtype=str), aba

def selecionar_coluna_sidebar(df: pd.DataFrame, chave: str, coluna_detectada: Optional[str]) -> Optional[str]:
    opcoes = ["-- selecione --"] + list(df.columns)
    idx = opcoes.index(coluna_detectada) if coluna_detectada in df.columns else 0
    escolha = st.sidebar.selectbox(COLUNAS_OBRIGATORIAS[chave]["rotulo"], opcoes, index=idx, key=f"col_{chave}")
    return None if escolha == "-- selecione --" else escolha

def calcular_icms_tdd(df_original, mapa_colunas, divisor_base, aliquota_pct, aplicar_aliquota, remover_linhas_total):
    df_base = df_original.copy()
    linhas_total = identificar_linhas_total(df_base)
    total_rem = int(linhas_total.sum()) if remover_linhas_total else 0
    if remover_linhas_total:
        df_base = df_base.loc[~linhas_total].copy()
    nums = {}
    for chave, coluna in mapa_colunas.items():
        nums[chave] = df_base[coluna].apply(converter_numero).astype(float)
    ali_dec = aliquota_pct / 100
    vlr_ad = nums["valor_total"] - nums["valor_ii"]
    soma = (vlr_ad + nums["valor_ii"] + nums["valor_ipi"] + nums["valor_pis"]
            + nums["valor_cofins"] + nums["siscomex"] + nums["afrmm"])
    base = soma / divisor_base
    if aplicar_aliquota:
        valor = base * ali_dec
        formula = f"Base ICMS TDD x {aliquota_pct:.4f}%"
    else:
        valor = base / ali_dec
        formula = f"Base ICMS TDD / {aliquota_pct:.4f}%"
    df_res = df_base.copy()
    df_res["Vlr Aduaneiro TDD"] = vlr_ad.round(2)
    df_res["Soma Base TDD"] = soma.round(2)
    df_res["Base ICMS TDD"] = base.round(2)
    df_res["Aliquota ICMS TDD %"] = aliquota_pct
    df_res["Valor ICMS TDD"] = valor.round(2)
    conf = localizar_colunas_conferencia(df_base)
    if conf.get("base_icms_original"):
        bo = df_base[conf["base_icms_original"]].apply(converter_numero).astype(float)
        df_res["Dif. Base ICMS TDD x Planilha"] = (base - bo).round(2)
    if conf.get("valor_icms_original"):
        vo = df_base[conf["valor_icms_original"]].apply(converter_numero).astype(float)
        df_res["Dif. ICMS TDD x Planilha"] = (valor - vo).round(2)
    totais = {"total_vlr_aduaneiro": float(vlr_ad.sum()), "total_soma_base": float(soma.sum()),
              "total_base_icms_tdd": float(base.sum()), "total_valor_icms_tdd": float(valor.sum()),
              "formula_final": formula}
    resumo = pd.DataFrame([
        {"Indicador": "Vlr Aduaneiro TDD", "Valor": totais["total_vlr_aduaneiro"]},
        {"Indicador": "Soma Base TDD", "Valor": totais["total_soma_base"]},
        {"Indicador": "Base ICMS TDD", "Valor": totais["total_base_icms_tdd"]},
        {"Indicador": "Valor ICMS TDD", "Valor": totais["total_valor_icms_tdd"]},
    ])
    return df_res, resumo, totais, total_rem

def gerar_excel_resultado(df_resultado: pd.DataFrame, df_resumo: pd.DataFrame) -> bytes:
    saida = io.BytesIO()
    with pd.ExcelWriter(saida, engine="openpyxl") as writer:
        df_resultado.to_excel(writer, index=False, sheet_name="Calculo_ICMS_TDD")
        df_resumo.to_excel(writer, index=False, sheet_name="Resumo")
        wb = writer.book
        hf = PatternFill("solid", fgColor="1F4E78")
        hfont = Font(color="FFFFFF", bold=True)
        for sn in ["Calculo_ICMS_TDD", "Resumo"]:
            ws = wb[sn]
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            for cell in ws[1]:
                cell.fill = hf
                cell.font = hfont
                cell.alignment = Alignment(horizontal="center", vertical="center")
            for ci, col in enumerate(ws.columns, start=1):
                nome = ws.cell(row=1, column=ci).value
                larg = min(max(12, max(len(str(c.value or "")) for c in col) + 2), 45)
                ws.column_dimensions[get_column_letter(ci)].width = larg
                if nome and any(t in str(nome).lower() for t in ["valor", "base", "soma", "aduaneiro", "dif."]):
                    for cel in ws.iter_cols(min_col=ci, max_col=ci, min_row=2, max_row=ws.max_row):
                        for item in cel:
                            item.number_format = '#,##0.00'
    saida.seek(0)
    return saida.getvalue()

def render_icms_tab() -> None:
    st.title("Gerar ICMS ALBEMA - TDD 409")
    st.caption("Importe a planilha da DUIMP/DI para calcular e conferir o ICMS do Sul.")
    with st.expander("Passo a passo", expanded=True):
        st.markdown(f"""
1. Acesse o portal [SAT/SEF-SC]({URL_SAT_SC})
2. Informe a DUIMP ou DI e solicite o ID no F5
3. Extraia a planilha Excel e importe aqui
4. No portal, altere o tratamento tributário para **TDD 409**
5. Compare os valores calculados com o portal
6. Exporte o arquivo final
        """)
    arquivo = st.file_uploader("Importe a planilha Excel ou CSV", type=["xlsx", "xls", "csv"], key="icms_up")
    if arquivo is None:
        return st.info("Aguardando importação da planilha.")
    try:
        df, origem = ler_arquivo_excel_ou_csv(arquivo.name, arquivo.getvalue())
    except Exception as e:
        return st.error(f"Erro ao ler planilha: {e}")
    if df.empty:
        return st.warning("Planilha vazia.")
    st.sidebar.header("Configuração")
    div = st.sidebar.number_input("Divisor da base ICMS", min_value=0.0001, value=0.96, step=0.01, format="%.4f")
    ali = st.sidebar.number_input("Alíquota ICMS TDD (%)", min_value=0.0001, value=2.60, step=0.10, format="%.4f")
    op = st.sidebar.radio("Operação", [
        "Aplicar alíquota: Base ICMS x 2,6%",
        "Dividir pela alíquota: Base ICMS / 2,6%",
    ], index=0)
    aplicar = op.startswith("Aplicar")
    remover = st.sidebar.checkbox("Remover linha de total", value=True)
    det = localizar_colunas_obrigatorias(df)
    mapa = {}
    for chave, col in det.items():
        escolha = selecionar_coluna_sidebar(df, chave, col)
        if escolha is not None:
            mapa[chave] = escolha
    falt = [COLUNAS_OBRIGATORIAS[c]["rotulo"] for c in COLUNAS_OBRIGATORIAS if c not in mapa]
    if falt:
        st.error(f"Colunas faltando: {', '.join(falt)}")
        return st.dataframe(df.head(20))
    df_res, df_sum, tots, rem = calcular_icms_tdd(df, mapa, div, ali, aplicar, remover)
    st.subheader("Resumo")
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Vlr Aduaneiro TDD", formatar_brl(tots["total_vlr_aduaneiro"]))
    c2.metric("Soma Base TDD", formatar_brl(tots["total_soma_base"]))
    c3.metric("Base ICMS TDD", formatar_brl(tots["total_base_icms_tdd"]))
    c4.metric("Valor ICMS TDD", formatar_brl(tots["total_valor_icms_tdd"]))
    st.write(f"**Origem:** {origem} | **Fórmula:** {tots['formula_final']}" + (f" | Linhas total removidas: {rem}" if rem else ""))
    st.dataframe(df_res, use_container_width=True)
    excel = gerar_excel_resultado(df_res, df_sum)
    st.download_button("Baixar Excel", data=excel, file_name="calculo_icms_tdd_albema.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")